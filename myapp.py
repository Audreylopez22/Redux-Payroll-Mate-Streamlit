import streamlit as st
import openpyxl

def main():
    st.title("Aquí puedes cargar tu archivo de excel")

    uploaded_file = st.file_uploader("Cargar un archivo Excel", type=["xlsx", "xls"])

    if uploaded_file is not None:
        workbook = openpyxl.load_workbook(uploaded_file)
        sheet = workbook.active

        if sheet.max_row == 1 and sheet.max_column == 1 and sheet.cell(row=1, column=1).value is None:
            st.warning("El archivo Excel está vacío.")
        else:
            st.success("El archivo Excel no está vacío.")

if __name__ == "__main__":
    main()
