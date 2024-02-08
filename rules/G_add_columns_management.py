from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def add_columns_management(sheet):
    target_sheets = {"Comp Management", "Guarapo"}

    if sheet.title in target_sheets:
        new_columns = [
            "Non cash out benefits",
            "Designated Cash out benefits",
            "Prorating",
            "On Going Reimbursements",
            "Sub Total",
            "Comments",
            "Bonus/Additional",
            "Total",
        ]

        last_column_number_before = sheet.max_column

        for i, column_name in enumerate(new_columns):
            new_column_number = last_column_number_before + i + 1
            new_column_letter = get_column_letter(new_column_number)
            sheet[f"{new_column_letter}1"] = column_name

        no_fill = PatternFill(fill_type="none")

        for col in range(last_column_number_before + 1, last_column_number_before + 8):
            for row in range(1, sheet.max_row + 1):
                sheet.cell(row=row, column=col).fill = no_fill

    return sheet


def main(workbook, progress_bar):
    for sheet in workbook:
        add_columns_management(sheet)

    if progress_bar is not None:
        progress_bar.progress(1.0 / len(sheet.parent.sheetnames))

    return workbook
