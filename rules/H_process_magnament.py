from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
import calendar
from datetime import datetime


def get_or_create_money_style(workbook):
    money_style_name = "money"

    # Check if the 'money' style already exists
    for style in workbook._named_styles:
        if style.name == money_style_name:
            return style

    # If it does not exist, create and add the style 'money'.
    money_style = NamedStyle(name=money_style_name, number_format='"$"#,##0.00')
    workbook.add_named_style(money_style)

    return money_style


def calculate_Prorating(monthly_salary, start_date):
    start_date = datetime.strptime(start_date, "%Y-%m-%d")
    start_day = start_date.day
    month = start_date.month
    year = start_date.year

    # Get the total number of days in the month, this returns the position of the first day of the month and number of days in the month
    total_days_in_month = calendar.monthrange(year, month)

    # Get the working days of the month (Monday to Friday)
    working_days_in_month = [
        day
        for day in range(1, total_days_in_month[1] + 1)
        if datetime(year, month, day).weekday() < 5
    ]
    # st.write(working_days_in_month)

    # Filter the working days from the start date to the end of the month
    working_days_worked = [day for day in working_days_in_month if day >= start_day]
    # st.write(working_days_worked)

    # Calculate the proportional salary
    total_working_days = len(working_days_in_month)
    worked_days = len(working_days_worked)

    proportional_salary = (monthly_salary / total_working_days) * worked_days

    return proportional_salary


def process_magnament(sheet):
    target_sheets = {"Comp Management", "Guarapo"}
    if sheet.title in target_sheets:
        last_row = sheet.max_row

        money_style = get_or_create_money_style(sheet.parent)
        now = datetime.now()
        current_year = now.year
        current_month = now.month

        for row in range(2, last_row + 1):
            monthly_salary = sheet[f"H{row}"].value
            start_date = sheet[f"I{row}"].value

            if (
                monthly_salary is not None
                and (start_date.month == current_month)
                and (start_date.year == current_year)
            ):
                proportional_salary = calculate_Prorating(
                    monthly_salary, start_date.strftime("%Y-%m-%d")
                )
                sheet[f"V{row}"].value = proportional_salary
                sheet[f"V{row}"].style = money_style
                print(f"Row {row}: proportional_salary={proportional_salary}")

            Q = get_column_letter(sheet["Q"][0].column)
            S = get_column_letter(sheet["S"][0].column)
            N = get_column_letter(sheet["N"][0].column)
            O = get_column_letter(sheet["O"][0].column)  # noqa: E741
            L = get_column_letter(sheet["L"][0].column)
            U = get_column_letter(sheet["U"][0].column)
            W = get_column_letter(sheet["W"][0].column)
            H = get_column_letter(sheet["H"][0].column)
            V = get_column_letter(sheet["V"][0].column)
            X = get_column_letter(sheet["X"][0].column)
            Z = get_column_letter(sheet["Z"][0].column)

            # Calculate Non cash out benefits
            sheet[f"T{row}"].value = f"={S}{row}+{Q}{row}"
            sheet[f"T{row}"].style = money_style

            # Designated Cash out benefits
            sheet[f"U{row}"].value = f"={N}{row}+{O}{row}+{L}{row}"
            sheet[f"U{row}"].style = money_style

            # On goin
            sheet[f"W{row}"].value = f"={U}{row}"
            sheet[f"W{row}"].style = money_style

            # Sub Total
            if (start_date.month == current_month) and (start_date.year == current_year):
                sheet[f"X{row}"].value = f"={V}{row}+{W}{row}"
                sheet[f"X{row}"].style = money_style

            else:
                sheet[f"X{row}"].value = f"={W}{row}+{H}{row}"
                sheet[f"X{row}"].style = money_style

            # Total
            sheet[f"AA{row}"].value = f"={X}{row}+ {Z}{row}"
            sheet[f"AA{row}"].style = money_style

        return sheet


def main(workbook, progress_bar):
    for sheet in workbook:
        process_magnament(sheet)

    if progress_bar is not None:
        progress_bar.progress(1.0 / len(sheet.parent.sheetnames))

    return workbook
