from tools import log_message
from openpyxl.utils import get_column_letter


def create_simple_sheet(workbook):
    if "Simple Sheet" not in workbook.sheetnames:
        log_message("creating simple sheet")
        simple_sheet = workbook.create_sheet("Simple Sheet")

        new_columns = [
            "Last name, First name",
            "Base Salary",
            "On Going Reimbursements",
            "Sub Total",
            "Comments",
            "Bonus/Additional",
            "Total",
        ]

        for i, column_name in enumerate(new_columns, start=1):
            new_column_letter = get_column_letter(i)
            simple_sheet[f"{new_column_letter}1"] = column_name


def main(workbook, progress_bar):
    create_simple_sheet(workbook)

    if progress_bar is not None:
        progress_bar.progress(1.0)

    return workbook
