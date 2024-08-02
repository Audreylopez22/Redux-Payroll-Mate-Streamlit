from tools import log_message
from openpyxl.utils import get_column_letter


def create_forecasting(workbook):
    if "Forecasting comp" not in workbook.sheetnames:
        log_message("creating forecasting comp")
        simple_sheet = workbook.create_sheet("Forecasting comp")

        new_columns = [
            "Last name, First name",
            "Division",
            "Pay rate",
            "Hire Date",
            "3 %",
            "total 3%",
            "5 %",
            "total 5%",
            "7%",
            "total 7%",
            "10%",
            "total 10 %",
        ]

        for i, column_name in enumerate(new_columns, start=1):
            new_column_letter = get_column_letter(i)
            simple_sheet[f"{new_column_letter}1"] = column_name


def main(workbook, progress_bar):
    create_forecasting(workbook)

    if progress_bar is not None:
        progress_bar.progress(1.0)

    return workbook
