from tools import log_message
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import streamlit as st
from tools import get_or_create_money_style

def create_guarapo(workbook):
   if "Guarapo" not in workbook.sheetnames:
        log_message("creating Guarapo")
        guarapo = workbook.create_sheet("Guarapo")

        new_columns = ["Last name, First name", "Country",
                                                "Status","Division","Job Title","Location",
                                                "Paid per","Pay rate","Hire Date","Birth Date",
                                                "Private Health Care Cloud Team Status",
                                                "Private Health Care Cloud Team Company pays",
                                                "Private Health Care Managed Teams Status",
                                                "Private Health Care Managed Teams Company pays",
                                                "Reimbursement for software license Company pays",
                                                "A Cloud Guru 2023 Effective date",
                                                "A Cloud Guru 2023 Company pays", 
                                                "Accounting Advise Status",
                                                "Accounting Advise Company pays",
                                                "Non cash out benefits", 
                                                "Designated Cash out benefits",
                                                "Prorating","On Going Reimbursements", 
                                                "Sub Total","Comments","Bonus/Additional", "Total"]

        for i, column_name in enumerate(new_columns, start=1):
            new_column_letter = get_column_letter(i)
            guarapo[f"{new_column_letter}1"] = column_name
            
def filter_and_display_data(sheet):
    
    if "Comp Management" == sheet.title:
        log_message(f"Filtering and displaying data for sheet: {sheet.title}")
        office_column_index = None
        header_row = sheet[1]
        for idx, cell in enumerate(header_row, start=1):
            if cell.value == "Location":
                office_column_index = idx

        if office_column_index is None:
            st.error("Column 'Location' not found.")
            return

        data_list = []
        for row in sheet.iter_rows(min_row=2, max_col=27, values_only=True): 
            office = row[office_column_index - 1]
            if office == "Guarapo B/quilla Oficce":
                data_list.append(row)

        if data_list:
            for row_data in data_list:
                sheet.parent["Guarapo"].append(row_data)

        else:
            st.warning("No data matching the criteria.")
            
def apply_money_style_to_range(sheet, start_col, end_col, money_style):
    for row in range(2, sheet.max_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.style = money_style

def process_guarapo(sheet):
    if sheet.title == "Guarapo":
        last_row= sheet.max_row
        
        money_style = get_or_create_money_style(sheet.parent)
        apply_money_style_to_range(sheet, 16, 27, money_style)
            
        for row in range (2, last_row +1):
            Q = 'Q'
            S = 'S'
            N = 'N'
            O = 'O'
            L = 'L'
            U = 'U'
            W = 'W'
            H = 'H'
            X = 'X'
            Z = 'Z'

            # Calculate Non cash out benefits
            sheet[f'T{row}'].value = f'={S}{row}'
            sheet[f'T{row}'].style = money_style
            
            # Designated Cash out benefits 
            sheet[f'U{row}'].value = f'={N}{row}+{O}{row}+{L}{row}+{Q}{row}'
            sheet[f'U{row}'].style = money_style
             
            # On goin 
            sheet[f'W{row}'].value = f'={U}{row}'
            sheet[f'W{row}'].style = money_style

            # Sub Total
            sheet[f'X{row}'].value = f'={W}{row}+{H}{row}'
            sheet[f'X{row}'].style = money_style
            
            # Total 
            sheet[f'AA{row}'].value = f'={X}{row}+ {Z}{row}'
            sheet[f'AA{row}'].style = money_style           

            
        return sheet
    
    return sheet


def main(workbook, progress_bar):
    for sheet in workbook:
        create_guarapo(workbook)
        filter_and_display_data(sheet) 
        
        if "Comp Management" == sheet.title:
            print(f"Filtered Data for sheet '{sheet.title}':")
            print(sheet.parent["Guarapo"].iter_rows(values_only=True))
        else:
            print(f"No data matching the criteria for sheet '{sheet.title}'.")
            
    for sheet in workbook:
        process_guarapo(sheet)
        
    
    if progress_bar is not None:
        progress_bar.progress(1.0 / len(sheet.parent.sheetnames))   

    return workbook
