from tools import log_message
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import streamlit as st
from tools import get_or_create_money_style

def create_guarapo(workbook):
   if "Guarapo" not in workbook.sheetnames:
        log_message("creating Guarapo")
        guarapo = workbook.create_sheet("Guarapo")

        new_columns = ["Last name, First name", "Country",
                                                "Status","Division","Job Title","Location",
                                                "Paid per","Pay rate","Hire Date","Birth Date",
                                                "Private Health Care Cloud Team Status",
                                                "Private Health Care Cloud Team Company pays",
                                                "Private Health Care Managed Teams Status",
                                                "Private Health Care Managed Teams Company pays",
                                                "Reimbursement for software license Company pays",
                                                "A Cloud Guru 2023 Effective date",
                                                "A Cloud Guru 2023 Company pays", 
                                                "Accounting Advise Status",
                                                "Accounting Advise Company pays"]

        for i, column_name in enumerate(new_columns, start=1):
            new_column_letter = get_column_letter(i)
            guarapo[f"{new_column_letter}1"] = column_name
            
def filter_and_display_data(sheet):
    
    if "Comp Management" == sheet.title:
        log_message(f"Filtering and displaying data for sheet: {sheet.title}")
        office_column_index = None
        header_row = sheet[1]
        for idx, cell in enumerate(header_row, start=1):
            if cell.value == "Location":
                office_column_index = idx

        if office_column_index is None:
            st.error("Column 'Location' not found.")
            return

        data_list = []
        for row in sheet.iter_rows(min_row=2, max_col=19, values_only=True): 
            office = row[office_column_index - 1]
            if office == "Guarapo B/quilla Oficce":
                data_list.append(row)

        if data_list:
            for row_data in data_list:
                sheet.parent["Guarapo"].append(row_data)

        else:
            st.warning("No data matching the criteria.")
            
def apply_money_style_to_range(sheet, start_col, end_col, money_style):
    for row in range(2, sheet.max_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.style = money_style

def main(workbook, progress_bar):

    create_guarapo(workbook)
    filter_and_display_data(workbook["Comp Management"]) 
        
    if "Comp Management" == workbook["Comp Management"]:
        print(f"Filtered Data for sheet 'Comp Management':")
        print(workbook["Guarapo"].iter_rows(values_only=True))
        
    
    if progress_bar is not None:
        progress_bar.progress(1.0)   

    return workbook
