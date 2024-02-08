from tools import log_message
import datetime
import streamlit as st
from openpyxl.styles import PatternFill


def find_row_index_by_name(sheet, name_to_find):
    # I search in the queue, taking into account the name
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2
    ):
        if row[0] == name_to_find:
            return row_index

    return None


def check_birthday_alert(sheet):
    if sheet.title != "Comp Management":
        return

    log_message(f"Checking birthday alert for sheet: {sheet.title}")
    birth_date_column_index = None
    header_row = sheet[1]
    for idx, cell in enumerate(header_row, start=1):
        if cell.value == "Birth Date":
            birth_date_column_index = idx

    if birth_date_column_index is None:
        st.error("No 'Birth Date' column found.")
        return

    # List of people with birthdays in the current month
    today = datetime.date.today()
    birthday_people = []
    for row in sheet.iter_rows(min_row=2, max_col=birth_date_column_index, values_only=True):
        birth_date = row[birth_date_column_index - 1]
        if isinstance(birth_date, datetime.date) and birth_date.month == today.month:
            birthday_people.append(row[0])
            name_to_find = row[0]

            row_index = find_row_index_by_name(sheet, name_to_find)
            # change of color of the birthday cell
            if row_index is not None:
                cell = sheet.cell(row=row_index, column=birth_date_column_index)
                cell.fill = PatternFill(
                    start_color="FFbdd7ee", end_color="FFbdd7ee", fill_type="solid"
                )

    if birthday_people:
        formatted_list = "\n".join([f"- {person}" for person in birthday_people])
        st.session_state.info.append(
            f"People with birthdays in the current month in sheet '{sheet.title}' : \n{formatted_list}"
        )

    else:
        st.warning("No birthdays in the current month'.")


def main(workbook, progress_bar):
    for sheet in workbook:
        check_birthday_alert(sheet)
    if progress_bar is not None:
        progress_bar.progress(1.0 / len(sheet.parent.sheetnames))

    return workbook
