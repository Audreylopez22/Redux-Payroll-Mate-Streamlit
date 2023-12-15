from tools import log_message
import streamlit as st
import pandas as pd
from openpyxl.styles import NamedStyle

def get_or_create_money_style(workbook):
    money_style_name = 'money'
    
    # # Check if the 'money' style already exists
    for style in workbook._named_styles:
        if style.name == money_style_name:
            return style

    #  If it does not exist, create and add the style 'money'.
    money_style = NamedStyle(name=money_style_name, number_format='"$"#,##0.00')
    workbook.add_named_style(money_style)
    
    return money_style

def assign_bonuses(workbook):
    log_message("Assigning bonuses from 'bonus sheet' to 'Comp Management'")
    
    # Select Sheets
    if "Bonus Sheet" in workbook.sheetnames:
        bonus_sheet = workbook["Bonus Sheet"]
        comp_management_sheet = workbook["Comp Management"]
        guarapo_sheet = workbook["Guarapo"]
        
        bonus_dict ={}
        
        money_style = get_or_create_money_style(workbook)
        
        for row in range (2, bonus_sheet.max_row + 1):
            name = bonus_sheet.cell(row=row, column=1).value
            bonus_amount = bonus_sheet.cell(row=row, column=11).value
            
            if bonus_amount is not None:
                if name in bonus_dict:
                    bonus_dict[name] += bonus_amount
                else:
                    bonus_dict[name] = bonus_amount
                
        for row in range(2, comp_management_sheet.max_row + 1):
            name_in_comp_management = comp_management_sheet.cell(row=row, column=1).value
            name_in_guarapo = guarapo_sheet.cell(row=row, column=1).value
            
            if name_in_comp_management in bonus_dict:
                bonus_value_comp_management = bonus_dict[name_in_comp_management]
                comp_management_sheet.cell(row=row, column=26, value=bonus_value_comp_management).style = money_style

            if name_in_guarapo in bonus_dict:
                bonus_value_guarapo = bonus_dict[name_in_guarapo]
                guarapo_sheet.cell(row=row, column=26, value=bonus_value_guarapo).style = money_style
                        
    return workbook

def main(workbook, progress_bar):

    assign_bonuses(workbook)
    
    if progress_bar is not None:
        progress_bar.progress(1.0)   

    return workbook
