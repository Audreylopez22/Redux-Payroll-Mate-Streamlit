
from tools import log_message
import datetime
import streamlit as st
from openpyxl.styles import PatternFill

def find_row_index_by_name(sheet, name_to_find, name_column_index):
    for row_index,row in enumerate(sheet.iter_rows(min_row=2, max_col=name_column_index,
                                                values_only=True), start=2):
        if row[0] == name_to_find:
            return row_index
    return None

def new_employee_alert(sheet):
    

    if sheet.title != "Comp Management":
        return

    log_message(f"checking for new employees this month for sheet: {sheet.title}")
    
    # Obtain the column of the date they entered the company
    hire_date_column_index = None
    header_row = sheet[1]
    for idx, cell in enumerate(header_row, start=1):
        if cell.value == "Hire Date":
            hire_date_column_index = idx
            break

    if hire_date_column_index is None:
        st.error("No 'Hire Date' column found.")
        return

    # Get a list of people who entered in the current month
    today = datetime.date.today()
    hireDate_people = []
    for row in sheet.iter_rows(min_row=2, max_col=hire_date_column_index, values_only=True):
        hire_date = row[hire_date_column_index - 1]
        if isinstance(hire_date, datetime.date) and hire_date.month == today.month and hire_date.year == today.year:
            hireDate_people.append(row[0]) 
            name_to_find = row[0]
            
            row_index = find_row_index_by_name(sheet, name_to_find,hire_date_column_index)

            if row_index is not None:
                cell = sheet.cell(row=row_index, column= hire_date_column_index)
                cell.fill = PatternFill(start_color="FFfee599", end_color="FFfee599", fill_type="solid") 

    if hireDate_people:
        formatted_list = "\n".join([f"- {person}" for person in hireDate_people])
        st.session_state.alerts.append(f"New employees in the current month in sheet '{sheet.title}' : \n{formatted_list}")

    else:
        st.session_state.info.append(f"No new employees were found in the current month.")
            
def main(workbook,progress_bar):
    for sheet in workbook:
        new_employee_alert(sheet)
    if progress_bar is not None:
        progress_bar.progress(1.0 / len(sheet.parent.sheetnames))

    return workbook
