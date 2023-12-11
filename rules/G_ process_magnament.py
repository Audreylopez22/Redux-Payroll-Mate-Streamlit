from tools import log_message
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle

def get_or_create_money_style(workbook):
    money_style_name = 'money'
    
    # Check if the 'money' style already exists
    for style in workbook._named_styles:
        if style.name == money_style_name:
            return style

    # If it does not exist, create and add the style 'money'.
    money_style = NamedStyle(name=money_style_name, number_format='"$"#,##0.00')
    workbook.add_named_style(money_style)
    
    return money_style

def process_magnament(sheet):
    if sheet.title == "Comp Management":
        last_row= sheet.max_row
        
        money_style = get_or_create_money_style(sheet.parent)
            
        for row in range (2, last_row +1):
            Q = get_column_letter(sheet['Q'][0].column)
            S = get_column_letter(sheet['S'][0].column)
            N = get_column_letter(sheet['N'][0].column)
            O = get_column_letter(sheet['O'][0].column)
            L = get_column_letter(sheet['L'][0].column)
            U = get_column_letter(sheet['U'][0].column)
            W = get_column_letter(sheet['W'][0].column)
            H = get_column_letter(sheet['H'][0].column)
            X = get_column_letter(sheet['X'][0].column)
            Z = get_column_letter(sheet['Z'][0].column)
            
            # Calculate Non cash out benefits
            sheet[f'T{row}'].value = f'={S}{row}'
            sheet[f'T{row}'].style = money_style
            
            # Designated Cash out benefits 
            sheet[f'U{row}'].value = f'={N}{row}+{O}{row}+{L}{row}+{Q}{row}'
            sheet[f'U{row}'].style = money_style
             
            # On goin 
            sheet[f'W{row}'].value = f'={U}{row}'
            sheet[f'W{row}'].style = money_style

            # Sub Total
            sheet[f'X{row}'].value = f'={W}{row}+{H}{row}'
            sheet[f'X{row}'].style = money_style
            
            # Total 
            sheet[f'AA{row}'].value = f'={X}{row}+ {Z}{row}'
            sheet[f'AA{row}'].style = money_style           
            
        return sheet
    
    return sheet

def main(workbook, progress_bar):
    
    for sheet in workbook:
        process_magnament(sheet)
    
    if progress_bar is not None:
        progress_bar.progress(1.0 / len(sheet.parent.sheetnames))   

    return workbook
