from tools import log_message, find_row_index_by_name
import datetime
import streamlit as st
from openpyxl.styles import PatternFill, NamedStyle


def get_or_create_money_style(workbook):
    money_style_name = "money"

    # # Check if the 'money' style already exists
    for style in workbook._named_styles:
        if style.name == money_style_name:
            return style

    # If it does not exist, create and add the style 'money'.
    money_style = NamedStyle(name=money_style_name, number_format='"$"#,##0.00')
    workbook.add_named_style(money_style)

    return money_style


def apply_money_style_to_range(sheet, start_col, end_col, money_style):
    for row in range(2, sheet.max_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.style = money_style


def process_forecasting(workbook):
    # Select Sheets
    if "Forecasting comp" in workbook.sheetnames:
        log_message("processing Forecasting comp")
        comp_management_sheet = workbook["Comp Management"]
        forecasting_comp = workbook["Forecasting comp"]

        forecasting_dic = {}

        date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")

        for row in range(2, comp_management_sheet.max_row + 1):
            name = comp_management_sheet.cell(row=row, column=1).value
            division = comp_management_sheet.cell(row=row, column=4).value
            pay_rate = comp_management_sheet.cell(row=row, column=8).value
            hire_date = comp_management_sheet.cell(row=row, column=9).value

            forecasting_dic[name] = [
                division,
                pay_rate,
                hire_date,
            ]

        forecasting_comp.delete_rows(2, forecasting_comp.max_row)

        for name, values in forecasting_dic.items():
            forecasting_comp.append([name] + values)

        for row in range(2, forecasting_comp.max_row + 1):
            hire_date_cell = forecasting_comp.cell(row=row, column=4)
            if isinstance(hire_date_cell.value, datetime.date):
                hire_date_cell.style = date_style

        today = datetime.date.today()
        next_month = today.month % 12 + 1
        for row in range(2, forecasting_comp.max_row + 1):
            name = forecasting_comp.cell(row=row, column=1).value
            hire_date = forecasting_dic.get(name, [None, None, None])[2]
            if hire_date and hire_date.month == next_month:
                forecasting_comp.cell(row=row, column=1).fill = PatternFill(
                    start_color="FFbdd7ee", end_color="FFbdd7ee", fill_type="solid"
                )

    return workbook


def calculate_forecasting(sheet):
    if sheet.title == "Forecasting comp":
        last_row = sheet.max_row

        money_style = get_or_create_money_style(sheet.parent)
        apply_money_style_to_range(sheet, 5, 9, money_style)

        for row in range(2, last_row + 1):
            # 3%
            sheet[f"E{row}"].value = f"=C{row}*$E$1"
            sheet[f"E{row}"].style = money_style

            # 5%
            sheet[f"F{row}"].value = f"=C{row}*$F$1"
            sheet[f"F{row}"].style = money_style

            # 7%
            sheet[f"G{row}"].value = f"=C{row}*$G$1"
            sheet[f"G{row}"].style = money_style

            # 10%
            sheet[f"H{row}"].value = f"=C{row}*$H$1"
            sheet[f"H{row}"].style = money_style

        return sheet

    return sheet


def check_aniversary_alert(sheet):
    if sheet.title != "Comp Management":
        return

    log_message(f"Checking anniversary alert for sheet: {sheet.title}")
    hire_date_column_index = None
    name_column_index = None
    header_row = sheet[1]
    for idx, cell in enumerate(header_row, start=1):
        if cell.value == "Hire Date":
            hire_date_column_index = idx
        elif cell.value == "Name":
            name_column_index = idx
            break

    if hire_date_column_index is None:
        st.error("No 'Hire Date' column found.")
        return

    today = datetime.date.today()
    next_month = today.month % 12 + 1
    hireDate_people = []
    for row in sheet.iter_rows(min_row=2, max_col=hire_date_column_index, values_only=True):
        hire_date = row[hire_date_column_index - 1]
        if isinstance(hire_date, datetime.date) and hire_date.month == next_month:
            hireDate_people.append(row[0])
            name_to_find = row[0]

            row_index = find_row_index_by_name(sheet, name_to_find, name_column_index)

            if row_index is not None:
                cell = sheet.cell(row=row_index, column=hire_date_column_index)
                cell.fill = PatternFill(
                    start_color="FFbdd7ee", end_color="FFbdd7ee", fill_type="solid"
                )

    if hireDate_people:
        formatted_list = "\n".join([f"- {person}" for person in hireDate_people])
        st.session_state.info.append(
            f"People with an anniversary in the next month in sheet '{sheet.title}' : \n{formatted_list}"
        )

    else:
        st.session_state.info.append("No anniversary in the next month")


def main(workbook, progress_bar):
    process_forecasting(workbook)
    calculate_forecasting(workbook["Forecasting comp"])

    for sheet in workbook:
        check_aniversary_alert(sheet)

    if progress_bar is not None:
        progress_bar.progress(1.0)

    return workbook
