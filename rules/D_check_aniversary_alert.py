from tools import log_message
import datetime
import streamlit as st
from openpyxl.styles import PatternFill


def find_row_index_by_name(sheet, name_to_find, name_column_index):
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=2, max_col=name_column_index, values_only=True), start=2
    ):
        if row[0] == name_to_find:
            return row_index
    return None


def check_aniversary_alert(sheet):
    if sheet.title != "Comp Management":
        return

    log_message(f"Checking aniversary alert for sheet: {sheet.title}")
    hire_date_column_index = None
    name_column_index = None
    header_row = sheet[1]
    for idx, cell in enumerate(header_row, start=1):
        if cell.value == "Hire Date":
            hire_date_column_index = idx
        elif cell.value == "Name":
            name_column_index = idx
            break

    if hire_date_column_index is None:
        st.error("No 'Hire Date' column found.")
        return

    today = datetime.date.today()
    hireDate_people = []
    for row in sheet.iter_rows(min_row=2, max_col=hire_date_column_index, values_only=True):
        hire_date = row[hire_date_column_index - 1]
        if isinstance(hire_date, datetime.date) and hire_date.month == today.month:
            hireDate_people.append(row[0])
            name_to_find = row[0]

            row_index = find_row_index_by_name(sheet, name_to_find, name_column_index)

            if row_index is not None:
                cell = sheet.cell(row=row_index, column=hire_date_column_index)
                cell.fill = PatternFill(
                    start_color="FFbdd7ee", end_color="FFbdd7ee", fill_type="solid"
                )

    if hireDate_people:
        formatted_list = "\n".join([f"- {person}" for person in hireDate_people])
        st.session_state.info.append(
            f"People with an aniversary in the current month in sheet '{sheet.title}' : \n{formatted_list}"
        )

    else:
        st.session_state.info.append("No aniversary in the current month")


def main(workbook, progress_bar):
    for sheet in workbook:
        check_aniversary_alert(sheet)
    if progress_bar is not None:
        progress_bar.progress(1.0 / len(sheet.parent.sheetnames))

    return workbook
