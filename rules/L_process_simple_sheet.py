from openpyxl.styles import NamedStyle


def get_or_create_money_style(workbook):
    money_style_name = "money"

    # # Check if the 'money' style already exists
    for style in workbook._named_styles:
        if style.name == money_style_name:
            return style

    # If it does not exist, create and add the style 'money'.
    money_style = NamedStyle(name=money_style_name, number_format='"$"#,##0.00')
    workbook.add_named_style(money_style)

    return money_style


def apply_money_style_to_range(sheet, start_col, end_col, money_style):
    for row in range(2, sheet.max_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.style = money_style


def process_simple_sheet(sheet):
    if sheet.title == "Simple Sheet":
        last_row = sheet.max_row

        money_style = get_or_create_money_style(sheet.parent)
        apply_money_style_to_range(sheet, 2, 6, money_style)

        for row in range(2, last_row + 1):
            B = "B"
            C = "C"
            D = "D"
            F = "F"

            # Calculate Sub Total
            sheet[f"D{row}"].value = f"={B}{row}+{C}{row}"
            sheet[f"G{row}"].style = money_style

            # Total
            sheet[f"G{row}"].value = f"={F}{row}+{D}{row}"
            sheet[f"G{row}"].style = money_style

        return sheet

    return sheet


def main(workbook, progress_bar):
    process_simple_sheet(workbook["Simple Sheet"])

    if progress_bar is not None:
        progress_bar.progress(1.0)

    return workbook
