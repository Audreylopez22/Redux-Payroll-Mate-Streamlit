from tools import log_message, find_row_index_by_name
import streamlit as st
from openpyxl.styles import PatternFill


def new_employee_alert(sheet):
    if sheet.title != "Comp Management":
        return
    log_message(f"checking inactive people this month for sheet: {sheet.title}")
    # get the column status
    status_column_index = None
    header_row = sheet[1]
    for idx, cell in enumerate(header_row, start=1):
        if cell.value == "Status":
            status_column_index = idx
            break

    if status_column_index is None:
        st.error("No 'Status' column found.")
        return

    # Get a list of people who are inactive
    status_inactive = "Inactive"
    status_people = []

    for row in sheet.iter_rows(min_row=2, max_col=status_column_index, values_only=True):
        status_column = row[status_column_index - 1]
        if isinstance(status_column, str) and status_column.lower() == status_inactive.lower():
            status_people.append(row[0])
            name_to_find = row[0]

            row_index = find_row_index_by_name(sheet, name_to_find, status_column_index)

            if row_index is not None:
                cell = sheet.cell(row=row_index, column=status_column_index)
                cell.fill = PatternFill(
                    start_color="FFfee599", end_color="FFfee599", fill_type="solid"
                )

    if status_people:
        formatted_list = "\n".join([f"- {person}" for person in status_people])
        st.session_state.alerts.append(
            f"Inactive employees in sheet '{sheet.title}': \n{formatted_list}"
        )

    else:
        st.session_state.info.append("No inactive employees were found in the current month.")


def main(workbook, progress_bar):
    for sheet in workbook:
        new_employee_alert(sheet)
    if progress_bar is not None:
        progress_bar.progress(1.0 / len(sheet.parent.sheetnames))

    return workbook
