from tools import log_message
from openpyxl.utils import get_column_letter

def process_magnament(sheet):
    if sheet.title == "Comp Management":
        last_row= sheet.max_row
        
        
        for row in range (2, last_row +1):
            R = sheet['R'][0].column
            T = sheet['T'][0].column
            O = sheet['O'][0].column
            P = sheet['P'][0].column
            X = sheet['X'][0].column
            I = sheet['I'][0].column
            V = sheet['V'][0].column
            Y = sheet['Y'][0].column
            U = sheet['U'][0].column
            
            # Calculate Non cash out benefits
            
            sheet[f'U{row}'].value = f'={get_column_letter(R)}{row}+{get_column_letter(T)}{row}'

            # Designated Cash out benefits 
            sheet[f'V{row}'].value = f'={get_column_letter(O)}{row}+{get_column_letter(P)}{row}'

            # Designated On Going Reimbursements 

            sheet[f'Y{row}'].value = f'={get_column_letter(X)}{row}+{get_column_letter(I)}{row}+{get_column_letter(V)}{row}'

            # Total 
 
            sheet[f'Z{row}'].value = f'={get_column_letter(Y)}{row}+ {get_column_letter(U)}{row}'

            
        return sheet
    
    return sheet

def main(workbook, progress_bar):
    for sheet in workbook:
        process_magnament(sheet)
    
    if progress_bar is not None:
        progress_bar.progress(1.0 / len(sheet.parent.sheetnames))   

    return workbook
