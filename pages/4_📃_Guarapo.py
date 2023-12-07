from tools import log_message
import streamlit as st
import io
from openpyxl import load_workbook
import pandas as pd
import os
import shutil

st.set_page_config(page_title="Guarapo", page_icon="📄", layout="wide")

st.markdown("# Guarapo")
st.sidebar.header("Guarapo")

def filter_and_display_data(sheet):
    log_message(f"Filtering and displaying data for sheet: {sheet.title}")

    if sheet.title == "COMP MANAGEMENT":
        office_column_index = None
        header_row = sheet[1]
        st.info("entra a la hoja")
        for idx, cell in enumerate(header_row, start=1):
            if cell.value == "Location":
                office_column_index = idx

        if office_column_index is None:
            st.error("Column 'location' not found.")
            return

        data_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            office = row[office_column_index - 1]
            if office == "Guarapo B/quilla Oficce":
                data_list.append(row)
        
        st.write(data_list)

        if data_list:
            df = pd.DataFrame(data_list, columns=[cell.value for cell in header_row])
            return df
        
        else:
            return pd.DataFrame()


def main():
        if 'tmp_file' not in st.session_state:
            st.warning("Cannot display data because no file has been uploaded.")
            return
        
        st.warning(st.session_state.tmp_file)
        """ st.warning(os.path.getsize(st.session_state.tmp_file)) """
        with open(st.session_state.tmp_file, "rb") as file_content:
            st.session_state.tmp_file_content = file_content.read() 

        uploaded_file = io.BytesIO(st.session_state.tmp_file_content)

        workbook = load_workbook(uploaded_file)
        
        if st.button("Exportar Archivo Original a Excel"):
            file_path = "uploaded_file.xlsx"

            # Guardar el contenido del archivo original en un nuevo archivo Excel
            with open(file_path, "wb") as original_file:
                original_file.write(st.session_state.tmp_file_content)

            st.success(f"Archivo original exportado como {file_path}")

        filtered_data = filter_and_display_data(workbook.active)

        st.write("Filtered Data:")
        st.write(filtered_data)
        
        if st.button("Export to Excel"):
            file_path = "filtered_data.xlsx"

            excel_data = io.BytesIO()
            filtered_data.to_excel(excel_data, index=False, header=True, engine="openpyxl")
            excel_data.seek(0)

            st.download_button(
                label="Download Excel File",
                key="download_filtered_data",
                data=excel_data,
                file_name=file_path,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.success(f"Data exported to {file_path}")
        
        """ if os.path.exists("/tmp"):
            files = os.listdir("/tmp")
            st.error(files)
            for file in files:
                if file.endswith(".xlsx"):
                    os.unlink(os.path.join(os.sep,"tmp",file))
            st.error(os.listdir("/tmp")) """
            
if __name__ == "__main__":
    main()
