import streamlit as st
import openpyxl
from io import BytesIO
import datetime
import importlib

st.set_page_config(page_title="Load Sheet", page_icon="📈", layout="wide")

if (
    "authentication_status" not in st.session_state
    or st.session_state.authentication_status is None
    or st.session_state.authentication_status is False
):
    st.warning("You must login to access this page.")
    st.markdown(
        f'<meta http-equiv="refresh" content="0;url={st.secrets.urls.login}">',
        unsafe_allow_html=True,
    )
    st.stop()

title = "Welcome to the Excel file processing application!"
st.markdown(f"# {title}")
st.sidebar.header(title)

st.write(
    """Here, you can upload your Excel file to perform various operations. Start by selecting your
    file via the 'Load an Excel file' button. The application will automatically check whether the
    file contains the necessary data. If it does, modifications will be made for the payroll
    validation process."""
)


def load_data_from_excel(uploaded_file):
    workbook = openpyxl.load_workbook(uploaded_file)
    sheet = workbook.active

    data = [[cell.value for cell in row] for row in sheet.iter_rows()]
    return data


def log_message(message):
    current_time = datetime.datetime.now()
    formatted_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
    st.text(f"[{formatted_time}] {message}")


def process_rules(workbook, progress_bar):
    try:
        # TERMINAR DE HACER ESTO
        rule_files = [
            "A_check_sheets.py",
            "B_validate_sheet.py",
            "C_check_birthday_alert.py",
            "D_check_aniversary_alert.py",
            "E_empty_cells.py",
            "F_guarapo.py",
            "G_add_columns_management.py",
            "H_process_magnament.py",
            "I_create_simple_sheet.py",
            "J_bonus.py",
            "K_simple_sheets_values.py",
            "L_distributing_payroll.py",
            "M_process_simple_sheet.py",
            "N_new_employee_alert.py",
            "O_alert_inactive_person.py",
            "P_forecasting.py",
            "Q_process_forecasting.py",
        ]
        total_steps = len(rule_files)

        for i, filename in enumerate(rule_files):
            rule_module_name = f"rules.{filename[:-3]}"
            rule = importlib.import_module(rule_module_name)
            if hasattr(rule, "main") and callable(rule.main):
                workbook = rule.main(workbook, progress_bar)

            # For the progres bar
            progress_percentage = min(1.0, (i + 1) / total_steps)
            progress_bar.progress(progress_percentage)

    except Exception as error:
        print(error)

    return workbook


def main():
    uploaded_file = st.file_uploader("Load an Excel file", type=["xlsx", "xls"])

    if uploaded_file is not None:
        # to reset the session state when analizyded new information
        st.session_state.alerts = []
        st.session_state.errors = []
        st.session_state.info = []
        st.session_state.logs = []
        uploaded_file_contents = uploaded_file.read()

        if "file_hash" not in st.session_state or st.session_state.file_hash != hash(
            uploaded_file_contents
        ):
            st.session_state.file_hash = hash(uploaded_file_contents)

        st.session_state.data = load_data_from_excel(uploaded_file)

        workbook = openpyxl.load_workbook(uploaded_file)
        sheet = workbook.active

        if (
            sheet.max_row == 1
            and sheet.max_column == 1
            and sheet.cell(row=1, column=1).value is None
        ):
            st.warning("The Excel file is empty.")
            log_message("The Excel file is empty.")

        else:
            progress_bar = st.progress(0.0)
            log_message("Making modifications to the Excel file...")

            workbook = process_rules(workbook, progress_bar)

            # the modified or processed file is saved in memory
            modified_file = BytesIO()
            log_message("Saving modifications to the Excel file...")
            workbook.save(modified_file)

            # download button for the downloaded file that is in memory
            st.session_state.download_button = st.download_button(
                label="Download modified file",
                data=modified_file.getvalue(),
                key="download_file.xlsx",
                file_name="modified_file.xlsx",
            )


if __name__ == "__main__":
    main()
